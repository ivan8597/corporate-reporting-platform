from pathlib import Path
import datetime
import pandas as pd

from openpyxl import Workbook, load_workbook

from utils.logger import get_logger


logger = get_logger(__name__)


def generate_excel_report(
    df: pd.DataFrame,
    kpis: dict,
    template_path="templates/report_template.xlsx"
):
    """
    Формирование Excel-отчёта без Microsoft Excel.
    Работает на Render/Linux.
    """

    logger.info(
        "Этап 4: Формирование профессионального Excel-отчёта"
    )

    output_dir = Path("data/reports")
    output_dir.mkdir(
        parents=True,
        exist_ok=True
    )

    date_str = datetime.datetime.now().strftime(
        "%Y%m%d_%H%M"
    )

    output_path = (
        output_dir /
        f"Корпоративный_Отчет_{date_str}.xlsx"
    )


    # Создаём Excel файл
    wb = Workbook()


    # Удаляем стандартный лист
    ws = wb.active
    ws.title = "Дашборд"


    # --------------------
    # Дашборд
    # --------------------

    ws["A1"] = "Показатель"
    ws["B1"] = "Значение"


    dashboard = [
        ("Выручка", kpis["Total_Revenue"]),
        ("Прибыль", kpis["Total_Profit"]),
        ("Средняя маржа %", kpis["Avg_Margin_%"]),
        ("Количество заказов", kpis["Total_Orders"]),
        ("Средний чек", kpis["Avg_Check"]),
    ]


    for row, item in enumerate(
        dashboard,
        start=2
    ):
        ws.cell(row=row, column=1).value = item[0]
        ws.cell(row=row, column=2).value = item[1]


    # --------------------
    # Продажи
    # --------------------

    sales_ws = wb.create_sheet(
        "Продажи"
    )


    for col, name in enumerate(
        df.columns,
        start=1
    ):
        sales_ws.cell(
            row=1,
            column=col
        ).value = name


    for row, data in enumerate(
        df.values,
        start=2
    ):
        for col, value in enumerate(
            data,
            start=1
        ):
            sales_ws.cell(
                row=row,
                column=col
            ).value = value


    # --------------------
    # ABC анализ
    # --------------------

    abc_ws = wb.create_sheet(
        "ABC-анализ"
    )


    abc = (
        df.groupby("product_name")["amount"]
        .sum()
        .reset_index()
        .sort_values(
            "amount",
            ascending=False
        )
    )


    abc["cum_percent"] = (
        abc["amount"]
        .cumsum()
        /
        abc["amount"].sum()
        *
        100
    )


    for col, name in enumerate(
        abc.columns,
        start=1
    ):
        abc_ws.cell(
            row=1,
            column=col
        ).value = name


    for row, data in enumerate(
        abc.values,
        start=2
    ):
        for col, value in enumerate(
            data,
            start=1
        ):
            abc_ws.cell(
                row=row,
                column=col
            ).value = value


    # --------------------
    # Ошибки данных
    # --------------------

    error_ws = wb.create_sheet(
        "Ошибки данных"
    )


    quality = df[
        df["data_quality"] != "OK"
    ]


    for col, name in enumerate(
        quality.columns,
        start=1
    ):
        error_ws.cell(
            row=1,
            column=col
        ).value = name


    for row, data in enumerate(
        quality.values,
        start=2
    ):
        for col, value in enumerate(
            data,
            start=1
        ):
            error_ws.cell(
                row=row,
                column=col
            ).value = value


    wb.save(output_path)


    logger.info(
        f"Excel-отчёт успешно сохранён: {output_path}"
    )


    return str(output_path)
